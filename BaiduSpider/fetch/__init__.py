"""
一个基于thread和queue的线程池，以任务为队列元素，动态创建线程，重复利用线程，
通过close和terminate方法关闭线程池。
"""
import queue
import threading
import contextlib

import urllib
import requests 
from bs4 import BeautifulSoup
import re
import time
import openpyxl
import os.path

 
# 创建空对象,用于停止线程
StopEvent = object()

timeArr = []
 
def callback(status, result):
    """
    根据需要进行的回调函数，默认不执行。
    :param status: action函数的执行状态
    :param result: action函数的返回值
    :return:
    """
    pass
 
 
def action(thread_name,company,index,total):
    """
    真实的任务定义在这个函数里
    :param thread_name: 执行该方法的线程名
    :param arg: 该函数需要的参数
    :return:
    """
#     try:
    print("%d 抓取%s公司的seo信息" %(index,company.logUsername))
    linkInfo = fetchRank(company) 
    linkInfo.setCompany(company)
    
#     write_to_excel(linkInfo,index)
    write_to_file(linkInfo,index)
    print("第%d个任务属于公司%s调用了线程 %s，并打印了这条信息！" % (index, company.logUsername, thread_name))
    if index==total:
        finishTime = time.clock()
        print("共耗时%f s"  %(finishTime-timeArr[0]))
#     except Exception as msg: 
#         print(msg)
             
 
class ThreadPool:
 
    def __init__(self, max_num, max_task_num=None):
        """
        初始化线程池
        :param max_num: 线程池最大线程数量
        :param max_task_num: 任务队列长度
        """
        # 如果提供了最大任务数的参数，则将队列的最大元素个数设置为这个值。
        if max_task_num:
            self.q = queue.Queue(max_task_num)
        # 默认队列可接受无限多个的任务
        else:
            self.q = queue.Queue()
        # 设置线程池最多可实例化的线程数
        self.max_num = max_num
        # 任务取消标识
        self.cancel = False
        # 任务中断标识
        self.terminal = False
        # 已实例化的线程列表
        self.generate_list = []
        # 处于空闲状态的线程列表
        self.free_list = []
 
    def put(self, func, args, callback=None):
        """
        往任务队列里放入一个任务
        :param func: 任务函数
        :param args: 任务函数所需参数
        :param callback: 任务执行失败或成功后执行的回调函数，回调函数有两个参数
        1、任务函数执行状态；2、任务函数返回值（默认为None，即：不执行回调函数）
        :return: 如果线程池已经终止，则返回True否则None
        """
        # 先判断标识，看看任务是否取消了
        if self.cancel:
            return
        # 如果没有空闲的线程，并且已创建的线程的数量小于预定义的最大线程数，则创建新线程。
        if len(self.free_list) == 0 and len(self.generate_list) < self.max_num:
            self.generate_thread()
        # 构造任务参数元组，分别是调用的函数，该函数的参数，回调函数。
        w = (func, args, callback,)
        # 将任务放入队列
        self.q.put(w)
 
    def generate_thread(self):
        """
        创建一个线程
        """
        # 每个线程都执行call方法
        t = threading.Thread(target=self.call)
        t.start()
 
    def call(self):
        """
        循环去获取任务函数并执行任务函数。在正常情况下，每个线程都保存生存状态，
        直到获取线程终止的flag。
        """
        # 获取当前线程的名字
        current_thread = threading.currentThread().getName()
        # 将当前线程的名字加入已实例化的线程列表中
        self.generate_list.append(current_thread)
        # 从任务队列中获取一个任务
        event = self.q.get()
        # 让获取的任务不是终止线程的标识对象时
        while event != StopEvent:
            # 解析任务中封装的三个参数
            func, arguments, callback = event
            # 抓取异常，防止线程因为异常退出
            company,index,total = arguments
            try:
                # 正常执行任务函数
                result = func(current_thread, *arguments)
                success = True
            except Exception as e:
                # 当任务执行过程中弹出异常
                result = None
                success = False
                print("thread exception %s %d" %(company.logUsername,index))
                print(e)
                log("error.txt", 'a', str(e)+"  "+company.logUsername+"  "+str(index))
#                 self.terminate()
            # 如果有指定的回调函数
            if callback is not None:
                # 执行回调函数，并抓取异常
                try:
                    callback(success, result)
                except Exception as e:
                    pass
            # 当某个线程正常执行完一个任务时，先执行worker_state方法
            with self.worker_state(self.free_list, current_thread):
                # 如果强制关闭线程的flag开启，则传入一个StopEvent元素
                if self.terminal:
                    event = StopEvent
                # 否则获取一个正常的任务，并回调worker_state方法的yield语句
                else:
                    # 从这里开始又是一个正常的任务循环
                    event = self.q.get()
        else:
            # 一旦发现任务是个终止线程的标识元素，将线程从已创建线程列表中删除
            self.generate_list.remove(current_thread)
 
    def close(self):
        """
        执行完所有的任务后，让所有线程都停止的方法
        """
        # 设置flag
        self.cancel = True
        # 计算已创建线程列表中线程的个数，然后往任务队列里推送相同数量的终止线程的标识元素
        full_size = len(self.generate_list)
        while full_size:
            self.q.put(StopEvent)
            full_size -= 1
 
    def terminate(self):
        """
        在任务执行过程中，终止线程，提前退出。
        """
        self.terminal = True
        # 强制性的停止线程
        while self.generate_list:
            self.q.put(StopEvent)
 
    # 该装饰器用于上下文管理
    @contextlib.contextmanager
    def worker_state(self, state_list, worker_thread):
        """
        用于记录空闲的线程，或从空闲列表中取出线程处理任务
        """
        # 将当前线程，添加到空闲线程列表中
        state_list.append(worker_thread)
        # 捕获异常
        try:
            # 在此等待
            yield
        finally:
            # 将线程从空闲列表中移除
            state_list.remove(worker_thread)


index_pattern = re.compile('http(s)?://(\w)*.cn.made-in-china.com(/)?')
prod_list_pattern = re.compile('.*/showroom/(\w)*-product-list-(\d)*.html')
prod_detail_pattern = re.compile('.*/gongying/.*')
video_pattern = re.compile('.*/video/.*')
gif_pattern = re.compile('.*/gif/.*')
tupian_pattern = re.compile('.*/tupian/.*')
qp_list_pattern = re.compile('.*(/hot-search/|-chanpin-|/cp/).*')
qc_list_pattern = re.compile('.*-gongsi-.*')
jiage_list_pattern = re.compile('.*/jiage/.*')
photo_list_pattern = re.compile('.*/photo/.*')


def getUrlType(url):    
    urlType = 'other'
    if index_pattern.match(url):
        urlType = 'index'
    elif prod_list_pattern.match(url):
        urlType = 'prod_list'
    elif prod_detail_pattern.match(url):
        urlType = 'prod_detail'
    elif tupian_pattern.match(url):
        urlType = 'tupian'   
    elif video_pattern.match(url):
        urlType = 'video'
    elif gif_pattern.match(url):
        urlType = 'gif' 
    elif qp_list_pattern.match(url):
        urlType = 'QP'     
    elif qc_list_pattern.match(url):
        urlType = 'QC' 
    elif jiage_list_pattern.match(url):
        urlType = 'jiage'     
    elif photo_list_pattern.match(url):
        urlType = 'photo'         
    return urlType    

class Company():
    def __init__(self,logUsername,comId,comName):
        self.logUsername = logUsername
        self.comId = comId
        self.comName = comName
        self.pageIndex = 0
        self.hasNext = True
class LinkInfo():
    def __init__(self,url,title,realTitle,urlType):
        self.url = url
        self.baiduTitle = title
        self.realTitle = realTitle
        self.urlType = urlType
    def getType(self):
        if self.urlType:
            return self.urlType
        else:
            self.type = getUrlType(self.urlType)
            return self.urlType

class SeoInfo():
    def __init__(self):
        self.linkInfoList = []
        self.company = None
    def setCompany(self,company):
        self.company = company    
    def addLink(self,linkInfo):    
        self.linkInfoList.append(linkInfo)
    def addLinks(self,linkInfoList):   
        self.linkInfoList.extend(linkInfoList)
        
class BaiduSpider():
    def __init__(self,keyword,pageIndex):
        self.keyword = keyword
        self.pageIndex = pageIndex
        self.searchUrl = ""
        self.pageContent = ""
        self.prefix = "http://www.baidu.com/s"
        pass
    
    def createUrl(self):
        command = "site:cn.made-in-china.com inurl:"+self.keyword
        search = [('wd', command), ('pn', self.pageIndex)]
        self.searchUrl = self.prefix + "?" + urllib.parse.urlencode(search)
        # print("searchUrl:%s" %(self.searchUrl))
    
    def getPageContent(self):
        self.createUrl()
        response = requests.get(url=self.searchUrl)
        self.pageContent = response.content
        return self.pageContent

class Analyzer():
    def __init__(self):
        pass
    def getSeoInfo(self,page):
        soup = BeautifulSoup(page,"html.parser")   
        tagh3 = soup.find_all('h3')
        div = soup.find_all('a',string = re.compile('下一页'))
        hasNext = True if div else False
        seoInfo = SeoInfo()
        for h3 in tagh3:
            href = h3.find('a').get('href')
            text = h3.get_text()
            res = requests.get(url=href)
            targetPage = res.content
            soup1 = BeautifulSoup(targetPage,"html.parser")   
            title = soup1.find_all('title')[0].get_text()
            real_url = res.request.url  #得到网页原始地址
            urlType = getUrlType(real_url)
            linkInfo = LinkInfo(real_url, text,title,urlType)
            seoInfo.addLink(linkInfo)
        return seoInfo,hasNext  

def log(fileName,model,msg):
    file_object = open(fileName, model,encoding='utf-8')
    file_object.write(msg+"\n")
    file_object.close() 
   
def creatwb(wbname):  
    wb=openpyxl.Workbook()
    wb.save(filename=wbname)
    print ("新建Excel："+wbname+"成功")

def write_to_file(seoInfo,index=0): 
    filename = 'seo.txt'
    excelArr = []
    com = seoInfo.company
    for linkInfo in seoInfo.linkInfoList:        
        excelArr.append([com.logUsername,com.comId,com.comName,linkInfo.urlType,linkInfo.url,linkInfo.baiduTitle,linkInfo.realTitle,str(len(seoInfo.linkInfoList))])
    
    for info in excelArr:
        value = "--|--".join(info)
        
        log(filename, 'a', value )
#     book.close()
    print("%d %s 写入 %s - 成功" %(index,com.logUsername,filename,))

        
def write_to_excel(seoInfo,index=0): 
    excel_index = index//500
    filename = 'links'+str(excel_index)+'.xlsx'
    sheetname = 'url_title'
    excelArr = []
    book = openpyxl.load_workbook(filename)  #打开excel文件   
    sheet=book[sheetname]
    
    com = seoInfo.company
    for linkInfo in seoInfo.linkInfoList:        
        excelArr.append([com.logUsername,com.comId,com.comName,linkInfo.urlType,linkInfo.url,linkInfo.baiduTitle,linkInfo.realTitle,len(seoInfo.linkInfoList)])
    
    max_row = sheet.max_row
    roww = 1 + max_row#控制行
    for info in excelArr:
        coll = 1#控制列
        for s in info:#再循环里面list的值，每一列
            log("access.txt", 'a', "log："+sheetname+"  " +str(max_row)+"  "+str(roww)+"  "+str(coll)+"\n" )
            sheet.cell(roww,coll,s)
            coll+=1
        roww+=1
    book.save(filename)#保存到当前目录下      
#     book.close()
    print("%d %s 写入 %s - %s成功" %(index,com.logUsername,filename,sheetname))

def read_from_excel(filename,startRowNum,endRowNum):
    wb=openpyxl.load_workbook(filename=filename,read_only=True)
    ws=wb.active
 
    data=[]
    for row in ws.iter_rows(min_row=startRowNum,max_row=endRowNum):
        row_info = []
        for cell in row:
            aa=str(cell.value)
            if (aa==""):
                aa="1"
            row_info.append(aa)
        company = Company(row_info[0],row_info[1],row_info[2])    
        data.append(company)
    print ("company.xlsx 读取成功！")
    return data    

def getExcelRowNum(filename):
    """
        获取Excel文件的行数
    :param filename: 文件名称    
    """
    wb=openpyxl.load_workbook(filename=filename,read_only=True)
    ws=wb.active
    return ws.max_row

def createExcel(num):
    excel_num = num//500
    for j in range(0,excel_num):
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = 'url_title'
        title = ['用户名','公司ID','公司名称','链接类型',' 收录链接 ',' 收录词 ' ,'实际title',' 收录数量 ']
        col = 1#控制列
        for s in title:#再循环里面list的值，每一列
            sheet.cell(1,col,s)
            col+=1
        book.save('links'+str(j)+'.xlsx')#保存到当前目录下    
        book.close()

def delExcel():
    if os.path.exists('seo.txt'):
        os.remove('seo.txt')
    if os.path.exists('error.txt'):
        os.remove('error.txt')
    if os.path.exists('access.txt'):
        os.remove('access.txt')
    if os.path.exists('links0.xlsx'):
        os.remove('links0.xlsx')
    if os.path.exists('links1.xlsx'):
        os.remove('links1.xlsx')
    if os.path.exists('links2.xlsx'):
        os.remove('links2.xlsx')
        
def fetchRank(company):
    seoInfo = SeoInfo()
    while(company.hasNext):
        baiuSpider = BaiduSpider(company.logUsername,company.pageIndex)     
        pageContent = baiuSpider.getPageContent()
        analyzer = Analyzer()
        tempSeoInfo,company.hasNext = analyzer.getSeoInfo(pageContent)
        seoInfo.addLinks(tempSeoInfo.linkInfoList)
        company.pageIndex = company.pageIndex + 10
    return seoInfo
# 调用方式
if __name__ == '__main__':
    timeArr.append(time.clock())     
 
    rowNum = getExcelRowNum("company.xlsx")
    print("共用%d条数据。" %(rowNum))
     
    delExcel()
    createExcel(rowNum)
      
    startRowNum = 1500
#     rowNum = 1500
    comInfos = read_from_excel("company.xlsx", startRowNum+1, rowNum+1);
    # 创建一个最多包含5个线程的线程池
    pool = ThreadPool(20,len(comInfos))
    # 创建100个任务，让线程池进行处理
    i=0
    for com in comInfos:
        i = i+1
        pool.put(action, (com,i,rowNum,), callback)
#         action('test',com,i,rowNum)
   
    
    # 强制关闭线程池
    # pool.terminate()