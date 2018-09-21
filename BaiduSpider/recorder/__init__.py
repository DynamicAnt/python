import openpyxl
import os.path
from _codecs import encode

info = []
def read_info_from_txt(filename):
    print("begin read")
    with open(filename,encoding='utf-8') as f:
        for line in f:
            print(line)
            if line =='\n':
                continue
            arr = line.split('--|--')
            info.append(arr)
    print("finish read")        

def write_info_to_excel(filename):
    if os.path.exists(filename):
        os.remove(filename)
    print("begin write")
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = 'url_title'
    title = ['用户名','公司ID','公司名称','链接类型',' 收录链接 ',' 收录词 ' ,'实际title',' 收录数量 ']
    col = 1#控制列
    for s in title:#再循环里面list的值，每一列
        sheet.cell(1,col,s)
        col+=1
        
    row = 2
    for seo in info:
        col = 1
        print(seo[0])
        for s in seo:#再循环里面list的值，每一列
            sheet.cell(row,col,s)
            col+=1 
        row+=1    
    book.save(filename)#保存到当前目录下 
    print("finish write")        

txt_name = "./seo/seo4.txt"
excel_name = './seo/links_txt4.xlsx'
read_info_from_txt(txt_name)
write_info_to_excel(excel_name)  