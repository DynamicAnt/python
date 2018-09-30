import openpyxl
import time
from db import mongodb


class ExcelUtil:
    @staticmethod
    def get_excel_row_num(filename):
        """
            获取Excel文件的行数
        :param filename: 文件名称
        """
        wb = openpyxl.load_workbook(filename=filename,read_only=True)
        ws = wb.active
        return ws.max_row

    @staticmethod
    def read_from_excel(filename, start, end):
        excel_start_time = time.clock()
        wb = openpyxl.load_workbook(filename=filename, read_only=True)
        ws = wb.active
        data = []
        for i in range(start, end):
            for j in range(12, 15):
                cell = ws.cell(row=i, column=j)
                value = cell.value
                if not value:
                    break
                else:
                    # if cell.value not in data:
                    data.append(cell.value)
        excel_end_time = time.clock()
        print("excel耗时：%f s" % (excel_end_time - excel_start_time))
        set_start_time = time.clock()
        temp = set(data)
        data = list(temp)
        set_end_time = time.clock()
        print("set耗时：%f s" % (set_end_time - set_start_time))
        return data


filename = ""
startTime = time.clock()
total = ExcelUtil.get_excel_row_num(filename) + 1
db_tool = mongodb.DB()
offset = 100
start = 2002
end = start + offset
while (end <= total) and (start != end):
    db_startTime = time.clock()
    words = ExcelUtil.read_from_excel(filename, start, end)
    db_tool.insert(words)
    print('已处理记录条数：%d' % end)
    start = end
    end = start + offset
    if end > total:
        end = total
    db_finishTime = time.clock()
    print("本次处理耗时：%f s" % (db_finishTime - db_startTime))

finishTime = time.clock()
print("共耗时%f s" % (finishTime-startTime))
# for value in word_lib:
#     print("value:" + value)
