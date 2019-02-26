"""
用于将mongodb 导出的json文件转成excel文件
"""
import openpyxl
import os.path
import json

info = []


def read_info_from_txt(filename):
    print("begin read")
    with open(filename, encoding='utf-8') as f:
        for line in f:
            print(line)

            if line == '\n':
                continue
            result = json.loads(line)
            info.append(result)
    print("finish read")


def write_info_to_excel(filename):
    if os.path.exists(filename):
        os.remove(filename)
    print("begin write")
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = 'url_title'
    title = ['keyword', 'ranking', 'username', 'com_name', 'cs_level', 'baidu_url', 'url', 'text']
    col = 1  # 控制列
    for s in title:  # 再循环里面list的值，每一列
        sheet.cell(1, col, s)
        col += 1

    row = 2
    for seo in info:
        col = 1
        print(seo['keyword'])
        for s in title:  # 再循环里面list的值，每一列
            sheet.cell(row, col, seo[s])
            col += 1
        row += 1
    book.save(filename)  # 保存到当前目录下
    print("finish write")


# txt_name = "./seo/seo4.txt"
# excel_name = './seo/links_txt4.xlsx'
txt_name = "E:\\python\\BaiduSpider\\utils\\seo\\seo1.json"
excel_name = 'E:\\python\\BaiduSpider\\utils\\seo\\seo1.xlsx'
read_info_from_txt(txt_name)
print("write excel")
write_info_to_excel(excel_name)

