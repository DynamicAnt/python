'''
Created on 2018年8月8日

@author: qiumingsheng
'''
#coding:utf-8
import urllib
from urllib import request
import requests 
from bs4 import BeautifulSoup
import re
import time
import openpyxl
import os.path


global pageIndex
pageIndex = 0

index_pattern = re.compile('http(s)?://(\w)*.cn.made-in-china.com(/)?')
prod_list_pattern = re.compile('.*/showroom/(\w)*-product-list-(\d)*.html')
prod_detail_pattern = re.compile('.*/gongying/.*')
video_pattern = re.compile('.*/video/.*')
gif_pattern = re.compile('.*/gif/.*')
tupian_pattern = re.compile('.*/tupian/.*')
qp_list_pattern = re.compile('.*(/hot-search/|-chanpin-|/cp/).*')
qc_list_pattern = re.compile('.*-gongsi-.*')
jiage_list_pattern = re.compile('.*/jiage/.*')
photo_list_pattern = re.compile('.*/photo/.*')


def getUrlType(url):    
    urlType = 'other'
    if index_pattern.match(url):
        urlType = 'index'
    elif prod_list_pattern.match(url):
        urlType = 'prod_list'
    elif prod_detail_pattern.match(url):
        urlType = 'prod_detail'
    elif tupian_pattern.match(url):
        urlType = 'tupian'   
    elif video_pattern.match(url):
        urlType = 'video'
    elif gif_pattern.match(url):
        urlType = 'gif' 
    elif qp_list_pattern.match(url):
        urlType = 'QP'     
    elif qc_list_pattern.match(url):
        urlType = 'QC' 
    elif jiage_list_pattern.match(url):
        urlType = 'jiage'     
    elif photo_list_pattern.match(url):
        urlType = 'photo'         
    return urlType    

class Company():
    def __init__(self,logUsername,comId,comName):
        self.logUsername = logUsername
        self.comId = comId
        self.comName = comName
        self.pageIndex = 0
        self.hasNext = True
class LinkInfo():
    def __init__(self,url,title,realTitle,urlType):
        self.url = url
        self.baiduTitle = title
        self.realTitle = realTitle
        self.urlType = urlType
    def getType(self):
        if self.urlType:
            return self.urlType
        else:
            self.type = getUrlType(self.urlType)
            return self.urlType

class SeoInfo():
    def __init__(self):
        self.linkInfoList = []
        self.company = None
    def setCompany(self,company):
        self.company = company    
    def addLink(self,linkInfo):    
        self.linkInfoList.append(linkInfo)
    def addLinks(self,linkInfoList):   
        self.linkInfoList.extend(linkInfoList)
        
class BaiduSpider():
    def __init__(self,keyword,pageIndex):
        self.keyword = keyword
        self.pageIndex = pageIndex
        self.searchUrl = ""
        self.pageContent = ""
        self.prefix = "http://www.baidu.com/s"
        pass
    
    def createUrl(self):
        command = "site:cn.made-in-china.com inurl:"+self.keyword
        search = [('wd',command),('pn',self.pageIndex)]
        self.searchUrl = self.prefix + "?" + urllib.parse.urlencode(search)
        print("searchUrl:%s" %(self.searchUrl))
    
    def getPageContent(self):
        self.createUrl()
        response = request.urlopen(self.searchUrl)
        page = response.read()
        self.pageContent = page.decode('utf-8')
        print("获取页面成功")
        return self.pageContent

class Analyzer():
    def __init__(self):
        pass
    def getSeoInfo(self,page):
        soup = BeautifulSoup(page,"html.parser")   
        tagh3 = soup.find_all('h3')
        div = soup.find_all('a',string = re.compile('下一页'))
        hasNext = True if div else False
        seoInfo = SeoInfo()
        for h3 in tagh3:
            href = h3.find('a').get('href')
            text = h3.get_text()
            res = requests.get(url=href)
            targetPage = res.content
            soup1 = BeautifulSoup(targetPage,"html.parser")   
            title = soup1.find_all('title')[0].get_text()
            real_url = res.request.url  #得到网页原始地址
            urlType = getUrlType(real_url)
            linkInfo = LinkInfo(real_url, text,title,urlType)
            seoInfo.addLink(linkInfo)
        print("finish getSeoInfo")  
        return seoInfo,hasNext  
   
def creatwb(wbname):  
    wb=openpyxl.Workbook()
    wb.save(filename=wbname)
    print ("新建Excel："+wbname+"成功")
        
def write_to_excel(seoInfo): 
    excelArr = []
    if not os.path.isfile('links.xlsx'):
        creatwb('links.xlsx')
        book = openpyxl.Workbook()#新建一个excel
        sheet = book.active
        sheet.title = 'url_title'
        excelArr.append(['用户名','公司ID','公司名称','链接类型',' 收录链接 ',' 收录词 ' ,'实际title',' 收录数量 '])
    else: 
        book = openpyxl.load_workbook('links.xlsx')  #打开excel文件   
        sheet=book['url_title']
      
    com = seoInfo.company
    for linkInfo in seoInfo.linkInfoList:        
        excelArr.append([com.logUsername,com.comId,com.comName,linkInfo.urlType,linkInfo.url,linkInfo.baiduTitle,linkInfo.realTitle,len(seoInfo.linkInfoList)])
    
    max_row = sheet.max_row
    row = 1#控制行
    for info in excelArr:
        col = 1#控制列
        for s in info:#再循环里面list的值，每一列
            sheet.cell(row+max_row,col,s)
            col+=1
        row+=1
    book.save('links.xlsx')#保存到当前目录下      
    print("finish write_to_excel")

def read_from_excel(filename,startRowNum,endRowNum):
    wb=openpyxl.load_workbook(filename=filename,read_only=True)
    ws=wb.active
 
    data=[]
    for row in ws.iter_rows(min_row=startRowNum,max_row=endRowNum):
        row_info = []
        for cell in row:
            aa=str(cell.value)
            if (aa==""):
                aa="1"
            row_info.append(aa)
        company = Company(row_info[0],row_info[1],row_info[2])    
        data.append(company)
    print ("company.xlsx 读取成功！")
    return data    

def getExcelRowNum(filename):
    """
        获取Excel文件的行数
    :param filename: 文件名称    
    """
    wb=openpyxl.load_workbook(filename=filename,read_only=True)
    ws=wb.active
    return ws.max_row

def fetchRank(company):
    seoInfo = SeoInfo()
    while(company.hasNext):
        baiuSpider = BaiduSpider(company.logUsername,company.pageIndex)     
        pageContent = baiuSpider.getPageContent()
        analyzer = Analyzer()
        tempSeoInfo,company.hasNext = analyzer.getSeoInfo(pageContent)
        seoInfo.addLinks(tempSeoInfo.linkInfoList)
        company.pageIndex = company.pageIndex + 10
#         if company.pageIndex > 20:
#             company.hasNext = False
    print("fetchRank finished")
    return seoInfo
 
 
 
startTime = time.clock()     

rowNum = getExcelRowNum("company.xlsx")
print("共用%d条数据。" %(rowNum))

startRowNum = 2
endRowNum =11
comInfos = read_from_excel("company.xlsx", startRowNum, endRowNum);

for comInfo in comInfos:
    print("抓取%s公司的seo信息" %(comInfo.logUsername))
    linkInfo = fetchRank(comInfo) 
    linkInfo.setCompany(comInfo)
    write_to_excel(linkInfo)
finishTime = time.clock()
    
print("共耗时%f s"  %(finishTime-startTime))